from openpyxl import * 
import warnings
import pdb

warnings.simplefilter("ignore")

def dayAnalysis(filename):
	if filename[-6] == '1':
		masterWorkbook1 = load_workbook("Day2Master.xlsx")
	if filename[-6] == '0':
		masterWorkbook1 = load_workbook("Day1Master.xlsx")

	masterWorksheet1 = masterWorkbook1.active

	#Opens the raw data file
	#filename = 'PR1D2_day_1.xlsx'
	dataWorkbook = load_workbook(filename) 
	dataWorksheet = dataWorkbook.active

	#WHAT IS EACH COUNTER FOR?
	#--> rollingNumber: iterate through the raw data file
	#--> familiar : first trial for both files should be 1234 and 4321; skip
	#--> masterKID: placeholder for location in output document
	rollingNumber = 1
	familiarization1 = False
	familiarization2 = False
	masterKID = 2

	blockCoordinate = "E" + str(rollingNumber)

	#Skip the first training block 

	while (familiarization1 == False or familiarization2 == False):
		if dataWorksheet.cell("E" + str(rollingNumber)).value == "BLOCK":
			if familiarization1 == False:
				familiarization1 = True
				random = True
			elif familiarization2 == False:
				familiarization2 = True
		rollingNumber +=1

	#After above code skips first training block, skip second training block

	while (dataWorksheet.cell("E" + str(rollingNumber)).value != "BLOCK"):
		rollingNumber +=1	

	#Start iterating through 3 blocks of learning dif seq
	#Initialize eRH to keep last response from each stim res pairing?

	blocksFound = 0
	extraResponseHour = None

	while blocksFound < 3:

		if (dataWorksheet.cell(blockCoordinate).value != "BLOCK"):
			rollingNumber +=1
			#Flip until you have first block

		#Found first block...
		elif (dataWorksheet.cell(blockCoordinate).value == "BLOCK" and familiarization1 == True and familiarization2 == True):

			blocksFound+=1
			if blocksFound != 1: 
				masterKID +=1
				masterWorksheet1["E" + str(masterKID)] = None

			rollingNumber += 1

			stimHour = None
			respHour = None

			sequenceCounter = 0 
			fromKey = 1000
			lastRespHour = 1

			#initialize fromKey to value that definitely won't meet conditions

			while sequenceCounter < 6:
				sequenceCounter+=1
				keyCounter = 0 
				while keyCounter < 10:

					while (dataWorksheet.cell("E" + str(rollingNumber)).value != "Stim 1") and (dataWorksheet.cell("E" + str(rollingNumber)).value != "Stim 2") and (dataWorksheet.cell("E" + str(rollingNumber)).value != "Stim 3") and (dataWorksheet.cell("E" + str(rollingNumber)).value != "Stim 4"): #and (dataWorksheet.cell("E" + str(rollingNumber)).value != "Blank"):
						rollingNumber+=1


					#CONDITION 1: THE STIMULUS IS BEFORE THE RESPONSE 
					
					if (dataWorksheet.cell("E" + str(rollingNumber)).value == "Stim 1" or "Stim 2" or "Stim 3" or "Stim 4"):

						stimHour = dataWorksheet.cell("A" + str(rollingNumber)).value
						stimMinute = dataWorksheet.cell("B" + str(rollingNumber)).value
						stimSec = dataWorksheet.cell("C" + str(rollingNumber)).value
						stimMsec = dataWorksheet.cell("D" + str(rollingNumber)).value

						rollingNumber +=1

						#Hit the stimulus...Condition 1 now is if it first has blank after

						if (dataWorksheet.cell("E" + str(rollingNumber)).value == "Blank"):
							rollingNumber +=1

						#Condition where there is no response needs to be right after the blank...
						
						if (dataWorksheet.cell("E" + str(rollingNumber)).value == "Stim 1") or (dataWorksheet.cell("E" + str(rollingNumber)).value ==  "Stim 2") or (dataWorksheet.cell("E" + str(rollingNumber)).value == "Stim 3") or  (dataWorksheet.cell("E" + str(rollingNumber)).value == "Stim 4"):
							# DO /NOT/ INCREMENT ROLLING NUMBER HERE!!!!!
							masterWorksheet1["E" + str(masterKID)] = None

						#Hit the stimulus...Condition 2 now is that it hits a response
						elif (dataWorksheet.cell("E" + str(rollingNumber)).value == "Response 1") or (dataWorksheet.cell("E" + str(rollingNumber)).value == "Response 2") or (dataWorksheet.cell("E" + str(rollingNumber)).value =="Response 3") or (dataWorksheet.cell("E" + str(rollingNumber)).value =="Response 4"):
							respHour = dataWorksheet.cell("A" + str(rollingNumber)).value
							respMinute = dataWorksheet.cell("B" + str(rollingNumber)).value
							respSec = dataWorksheet.cell("C" + str(rollingNumber)).value
							respMsec = dataWorksheet.cell("D" + str(rollingNumber)).value
							if dataWorksheet.cell("E" + str(rollingNumber)).value == "Response 1":
								responseGiven = 1
							elif dataWorksheet.cell("E" + str(rollingNumber)).value == "Response 2":
								responseGiven = 2
							elif dataWorksheet.cell("E" + str(rollingNumber)).value == "Response 3":
								responseGiven = 3
							elif dataWorksheet.cell("E" + str(rollingNumber)).value == "Response 4":
								responseGiven = 4
							rollingNumber +=1

						respSync1 = ((respHour - stimHour) * 3600000) + ((respMinute - stimMinute) * 60000) + ((respSec - stimSec) * 1000) + (respMsec - stimMsec)
						if ((type(lastRespHour) == int) and (fromKey == keyCounter -1)):
							respSync2 = ((stimHour - lastRespHour) * 3600000) + ((stimMinute - lastRespMinute) * 60000) + ((stimSec - lastRespSec) * 1000) + (stimMsec - lastRespMsec)
							if (-100 < respSync2 < 0):
								masterWorksheet1["E" + str(masterKID)] = respSync2
							else:
								masterWorksheet1["E" + str(masterKID)] = None

						elif -100 < respSync1 < 900:
							masterWorksheet1["E" + str(masterKID)] = respSync1
						else:
							masterWorksheet1["E" + str(masterKID)] = None

							"""RESPONSE SYNC 2 IF IT EXISTS
							   RESPONSE SYNC 2 ONLY IF IT FITS RULES
							   INCLUDE RULES FOR ALL
							"""
				
						masterWorksheet1["G" + str(masterKID)] = responseGiven
						if masterWorksheet1["G" + str(masterKID)].value == masterWorksheet1["C" + str(masterKID)].value:

							masterWorksheet1["F" + str(masterKID)] = 1
						else:

							masterWorksheet1["F" + str(masterKID)] = 0

						masterKID +=1


						while (dataWorksheet.cell("E" + str(rollingNumber)).value == "Response 1") and (dataWorksheet.cell("E" + str(rollingNumber)).value != "Response 2") and (dataWorksheet.cell("E" + str(rollingNumber)).value != "Response 3") and (dataWorksheet.cell("E" + str(rollingNumber)).value != "Response 4") and (dataWorksheet.cell("E" + str(rollingNumber)).value != "Blank"):
							print "More than 1 responses found :("
							if (dataWorksheet.cell("E" + str(rollingNumber)).value != "Blank"):
								lastRespHour = dataWorksheet.cell("A" + str(rollingNumber)).value
								lastRespMin = dataWorksheet.cell("B" + str(rollingNumber)).value
								lastRespSec = dataWorksheet.cell("C" + str(rollingNumber)).value
								lastRespMsec = dataWorksheet.cell("D" + str(rollingNumber)).value
								rollingNumber +=1
								fromKey = keyCounter 
							else: 
								rollingNumber+=1

						if (dataWorksheet.cell("E" + str(rollingNumber)).value == None):
							keyCounter = 11

					keyCounter+=1

	filenameSave = str(filename[0:-5]) + '_ACC.xlsx'
	masterWorkbook1.save(filenameSave)




