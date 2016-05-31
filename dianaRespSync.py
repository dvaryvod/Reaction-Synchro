#imports

from openpyxl import * 
import warnings

#root through each user's day 1 and day 2 data

def fileName(letter, ID, day):
	"""
	Arguments: letter: E or C (experiment/stutter or control)
	ID: participant ID 
	day: 1 or 2

	Returns: string filename of stats

	"""

	filename = "P" + letter + str(ID) + "D" + str(day) + "_day_" + str((day-1)) + "_ACC" + ".xlsx"
	return filename 

def calculateMeanAccuracy(worksheet, start):
	"""
	Arguments: worksheet: name of file with participant data (usually call to fileName)
	start: cell at which block starts (Should be 2,63,124...Expect TypeError otherwise)

	Returns: list of 6 values for accuracy for each of 6 trials in the block 
	"""
	i = start
	j = start
	meanAccuracies = []
	a = 1
	while a < 7:
		accuracyTempList = []
		while i < j+10:
			coordinate2 = "F" + str(i)
			value =  worksheet.cell(coordinate2).value
			accuracyTempList.append(int(value))
			i = i+1
		num = float(sum(accuracyTempList))
		den = float(len(accuracyTempList))
		mean = num/den
		meanAccuracies.append(mean)
		i = j + 10
		j = j+ 10 
		a = a+1
	return meanAccuracies

def calculateRespSync(worksheet, start):
	"""
	Same arguments/return as calculateMeanAccuracy except response synchrony

	Note that it only takes sequences where participant was accurate for all keypresses.
	If participant made mistakes, value for sequence resp synchrony will be None type 
	"""
	i = start
	j = start
	meanRespSync = []
	a = 1
	while a < 7:
		syncTempList = []
		while i < j+10:
			coordinate3 = "E" + str(i)
			value =  worksheet.cell(coordinate3).value
			try:
				syncTempList.append(int(value))
			except (TypeError): 
				syncTempList.append(None)
			finally:
				i = i+1
		try:
			num = float(sum(syncTempList))
			den = float(len(syncTempList))
			mean = num/den
			meanRespSync.append(mean)	
		except (TypeError):
			meanRespSync.append(None)
		finally:
			i = j + 10
			j = j+ 10 
			a = a+1
	return meanRespSync

#getStats('C', 1,1)

def main(filename):

	"""
	When indexing accuracyList and respSyncList: list[int] will give you list for that block of 6
	Note that you can change name of output file in .save() call
	Warnings suppressed

	Argument:
	filename: the file to open for data
	Returns: lists of accuracy, blocks and response synchrony
	"""			


#	filename = fileName('C', 1, 1)
	warnings.simplefilter("ignore")
	workbook = load_workbook(filename) 
	worksheet = workbook.active

	#Note: with current formatting, in column A, cells have value of None if 
	#it's from aforementioned sequence

	accuracyList = []
	respSyncList = []
	sequenceList = []
	blockList = [1,1,1,1,1,1,2,2,2,2,2,2,3,3,3,3,3,3]
	for i in range(2,184):
		coordinate = "A" + str(i)
		if worksheet.cell(coordinate).value != None:
			sequenceList.append(int(worksheet.cell(coordinate).value))

	accuracyList.append(calculateMeanAccuracy(worksheet, 2))
	accuracyList.append(calculateMeanAccuracy(worksheet, 63))
	accuracyList.append(calculateMeanAccuracy(worksheet, 124))

	respSyncList.append(calculateRespSync(worksheet,2))
	respSyncList.append(calculateRespSync(worksheet,63))
	respSyncList.append(calculateRespSync(worksheet,124))

	return accuracyList, respSyncList, blockList

	#counter1 = 2
	#counter2 = counter1+19
	#while counter1 < counter2:
	#	coordinate4 = "A" + str(counter1)
	#	activeSheet[coordinate4] = 

#	activeSheet["A2"] = "C"
#	activeSheet['B2'] = '1'
#	activeSheet['C2'] = blockList[0]
#	activeSheet['D2'] = sequenceList[0]
#	activeSheet['E2'] = meanAccuracies[0]
#	activeSheet['F2'] = respSyncList[0]

def coordinateMake(group, rowCounter):

	"""
	Arguments are group (C/E) and which row in Excel file this should be added to.
	Returns cell coordinates for Excel.
	"""

	coordinateGroup = "A" + str(rowCounter)
	coordinateID = "B" + str(rowCounter)
	coordinateBlock = "C" + str(rowCounter)
	coordinateSeq = "D" + str(rowCounter)
	coordinateAccuracy = "E" + str(rowCounter)
	coordinateResp = "F" + str(rowCounter)

	return coordinateGroup, coordinateID, coordinateBlock, coordinateSeq, coordinateAccuracy, coordinateResp

def saveFile():

	workbookSaveData = Workbook()
	workbookSaveData.active.title = "Accuracy and Resp Sync"
	activeSheet = workbookSaveData.get_sheet_by_name("Accuracy and Resp Sync")
	activeSheet["A1"] = "Participant Group"
	activeSheet["B1"] = "Participant ID"
	activeSheet["C1"] = "Block"
	activeSheet["D1"] = "Sequence Done"
	activeSheet["E1"] = "Accuracy %"
	activeSheet["F1"] = "Resp Sync"
	workbookSaveData.save('trialSave')

	studyGroups = ["C", "E"]
	controlID = [1, 3, 4, 6, 12, 13, 16, 17]
	stutterID = [2, 4, 5, 6, 7, 9, 10, 11]

	rowCounter = 2

	for group in studyGroups:

		if group == "C":

			for ID in controlID: 
				individualCounter = 0
				name1 = fileName(group, controlID, 1)
				main(name1)
				coordinateMake(group, rowCounter)
				rowCounter = rowCounter+1
				activeSheet[coordinateGroup] = group
				activeSheet[coordinateID] = ID
				activeSheet[coordinateBlock] = blockList[individualCounter]
				activeSheet[coordinateSeq] = sequenceList[individualCounter]
				activeSheet[coordinateAccuracy] = accuracyList[individualCounter]
				individualCounter = individualCounter +=1 

				name2 = fileName(group, controlID, 2)
				individualCounter = 0
				main(name2)
				coordinateMake(group, rowCounter)
				rowCounter = rowCounter+1
				activeSheet[coordinateGroup] = group
				activeSheet[coordinateID] = ID
				activeSheet[coordinateBlock] = blockList[individualCounter]
				activeSheet[coordinateSeq] = sequenceList[individualCounter]
				activeSheet[coordinateAccuracy] = accuracyList[individualCounter]
				individualCounter = individualCounter +=1 
		
		else: 

			for ID in stutterID: 
				name1 = fileName(group, stutterID, 1)
				main(name1)
				coordinateMake(group, rowCounter)
				rowCounter = rowCounter+1
				activeSheet[coordinateGroup] = group
				activeSheet[coordinateID] = ID
				activeSheet[coordinateBlock] = blockList[individualCounter]
				activeSheet[coordinateSeq] = sequenceList[individualCounter]
				activeSheet[coordinateAccuracy] = accuracyList[individualCounter]
				individualCounter = individualCounter +=1

				name2 = fileName(group, stutterID, 2)	
				main(name2)
				coordinateMake(group, rowCounter)
				rowCounter = rowCounter+1
				activeSheet[coordinateGroup] = group
				activeSheet[coordinateID] = ID
				activeSheet[coordinateBlock] = blockList[individualCounter]
				activeSheet[coordinateSeq] = sequenceList[individualCounter]
				activeSheet[coordinateAccuracy] = accuracyList[individualCounter]
				individualCounter = individualCounter +=1 

